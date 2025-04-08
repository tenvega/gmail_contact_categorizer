"""
Excel exporter for Gmail Contact Categorizer.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """Exports categorized contacts to Excel format."""
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.contacts = {}
    
    def export(self, categories, output):
        """
        Export categorized contacts to Excel.
        
        Args:
            categories (dict): Dictionary of categorized contacts.
            output (str): Output Excel filename.
            
        Returns:
            str: Path to the exported Excel file.
        """
        # Store contacts for data preparation
        self.contacts = categories
        
        # Prepare data for export
        export_data = self._prepare_export_data(categories)
        
        # Create DataFrame
        df = pd.DataFrame(export_data)
        
        # Create Excel writer
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Contacts', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Contacts']
        
        # Apply formatting
        self._apply_excel_formatting(worksheet)
        
        # Save the file
        writer.close()
        
        return output
    
    def _apply_excel_formatting(self, worksheet):
        """
        Apply formatting to the Excel worksheet.
        
        Args:
            worksheet: OpenPyXL worksheet object.
        """
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        cell_font = Font(name='Calibri')
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = cell_alignment
            cell.border = border
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Freeze panes
        worksheet.freeze_panes = 'A2'
        
        # Add filters
        worksheet.auto_filter.ref = worksheet.dimensions
